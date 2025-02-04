from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart,Reference


wb = load_workbook('barchart.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


# sheet['B8'] = 'SUM(B6:B7)'
# sheet['B8'].style = 'Currency'

# sheet['C8'] = 'SUM(C6:C7)'
# sheet['C8'].style = 'Currency'


for i in range(min_column+1,max_column+1):
    letter = get_column_letter(i)
    
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row+1})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

wb.save('report.xlsx')